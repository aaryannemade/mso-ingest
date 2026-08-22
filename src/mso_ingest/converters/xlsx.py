"""``.xlsx`` / ``.xls`` -> one CSV per worksheet, in workbook order."""

from __future__ import annotations

from pathlib import Path

from .. import detect, external
from ..manifest import CSV
from .base import Context, Result, reset_dir, sanitize, unique


def convert(ctx: Context) -> Result:
    result = Result()
    # Decided by content, not extension: legacy workbooks renamed to .xlsx are
    # common, and openpyxl cannot read them.
    is_legacy = detect.is_ole2(ctx.source)

    try:
        sheets = _sheets(ctx.source, legacy=is_legacy)
    except Exception as exc:  # noqa: BLE001 - surfaced as a per-document error
        result.warn(f"could not read workbook structure: {exc}", degraded=True)
        return result

    if not sheets:
        result.warn("workbook contains no worksheets", degraded=True)
        return result

    sheets_dir = reset_dir(ctx.outdir / "sheets")
    taken: set[str] = set()
    root = ctx.outdir.parent

    for index, (name, dims) in enumerate(sheets, start=1):
        # Sheet names allow characters that filenames do not, and two sheets
        # can sanitise to the same string, so uniqueness is enforced here.
        filename = unique(sanitize(name, fallback=f"sheet-{index}"), taken)
        dest = sheets_dir / f"{filename}.csv"

        try:
            external.in2csv_sheet(ctx.source, name, dest, fmt="xls" if is_legacy else None)
        except external.ToolError as exc:
            result.warn(f"sheet {name!r} failed to convert: {exc}", degraded=True)
            dest.unlink(missing_ok=True)
            continue

        result.add(
            dest,
            CSV,
            root=root,
            meta={"sheet": name, "index": index, "rows": dims[0], "columns": dims[1]},
        )

    return result


def _sheets(path: Path, *, legacy: bool) -> list[tuple[str, tuple[int, int]]]:
    """Return ``[(sheet_name, (rows, columns)), ...]`` in workbook order."""
    if legacy:
        import xlrd

        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            out = []
            for name in book.sheet_names():
                sheet = book.sheet_by_name(name)
                out.append((name, (sheet.nrows, sheet.ncols)))
                book.unload_sheet(name)
            return out
        finally:
            book.release_resources()

    import openpyxl

    # data_only=True yields the values Excel last computed rather than formula
    # source, which is what a downstream consumer of the CSV wants.
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            (name, (workbook[name].max_row or 0, workbook[name].max_column or 0))
            for name in workbook.sheetnames
        ]
    finally:
        workbook.close()
