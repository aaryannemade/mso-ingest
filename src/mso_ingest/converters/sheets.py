"""Read a workbook's structure without converting any of its data."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class Sheet:
    """One worksheet, as the manifest describes it."""

    name: str
    rows: int
    columns: int


def describe(path: Path, *, legacy: bool) -> list[Sheet]:
    """Return every worksheet in ``path``, in workbook order."""
    return _legacy_sheets(path) if legacy else _ooxml_sheets(path)


def _ooxml_sheets(path: Path) -> list[Sheet]:
    import openpyxl

    # data_only=True reports the values Excel last computed rather than the
    # formula source, which is what a consumer of the CSV wants.
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            Sheet(
                name=name,
                rows=workbook[name].max_row or 0,
                columns=workbook[name].max_column or 0,
            )
            for name in workbook.sheetnames
        ]
    finally:
        workbook.close()


def _legacy_sheets(path: Path) -> list[Sheet]:
    import xlrd

    book = xlrd.open_workbook(str(path), on_demand=True)
    try:
        found = []
        for name in book.sheet_names():
            sheet = book.sheet_by_name(name)
            found.append(Sheet(name=name, rows=sheet.nrows, columns=sheet.ncols))
            book.unload_sheet(name)
        return found
    finally:
        book.release_resources()
